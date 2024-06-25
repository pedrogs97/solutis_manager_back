"""Asset service"""

import logging
from io import BytesIO
from typing import List, Union

from fastapi import UploadFile, status
from fastapi.exceptions import HTTPException
from fastapi_pagination import Page, Params
from fastapi_pagination.ext.sqlalchemy import paginate
from openpyxl import load_workbook
from sqlalchemy import desc
from sqlalchemy.orm import Session

from src.asset.filters import AssetFilter, AssetStatusFilter, AssetTypeFilter
from src.asset.models import (
    AssetModel,
    AssetStatusHistoricModel,
    AssetStatusModel,
    AssetTypeModel,
)
from src.asset.schemas import (
    AssetSerializerSchema,
    AssetStatusSerializerSchema,
    AssetTypeSerializerSchema,
    InactivateAssetSchema,
    NewAssetSchema,
    UpdateAssetSchema,
)
from src.auth.models import UserModel
from src.config import DEFAULT_DATE_FORMAT
from src.invoice.models import InvoiceModel
from src.lending.models import LendingModel
from src.lending.schemas import (
    CostCenterSerializerSchema,
    LendingAssetHistorySerializerSchema,
)
from src.log.services import LogService
from src.people.schemas import EmployeeShortSerializerSchema

logger = logging.getLogger(__name__)
service_log = LogService()


class AssetService:
    """Asset services"""

    COL_NAMES = [
        "Patrimônio",
        "Descrição",
        "Fornecedor",
        "Garantia",
        "Observações",
        "Padrão",
        "Sistema Operacional",
        "N° Serial",
        "IMEI",
        "Data de Aquisição",
        "Valor",
        "Pacote Office",
        "Linha Telefônica",
        "Operadora",
        "Modelo",
        "Acessórios",
        "Configuração",
        "Quantidade",
        "Unidade",
        "Nota Fiscal",
        "Tipo de Ativo",
        "Status do Ativo",
    ]

    MAP_COL_NAMES = {
        "Patrimônio": "register_number",
        "Descrição": "description",
        "Fornecedor": "supplier",
        "Garantia": "assurance_date",
        "Observações": "observations",
        "Padrão": "pattern",
        "Sistema Operacional": "operational_system",
        "N° Serial": "serial_number",
        "IMEI": "imei",
        "Data de Aquisição": "acquisition_date",
        "Valor": "value",
        "Pacote Office": "ms_office",
        "Linha Telefônica": "line_number",
        "Operadora": "operator",
        "Modelo": "model",
        "Acessórios": "accessories",
        "Configuração": "configuration",
        "Quantidade": "quantity",
        "Unidade": "unit",
        "Nota Fiscal": "invoice_id",
        "Tipo de Ativo": "type_id",
        "Status do Ativo": "status_id",
    }

    def __get_asset_or_404(self, asset_id: int, db_session: Session) -> AssetModel:
        """Get asset or 404"""
        asset = db_session.query(AssetModel).filter(AssetModel.id == asset_id).first()
        if not asset:
            db_session.close()
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"field": "assetId", "error": "Ativo não encontrado"},
            )

        return asset

    def __validate_nested(
        self, data: Union[NewAssetSchema, UpdateAssetSchema], db_session: Session
    ) -> tuple:
        """Validates clothing size, type and status values"""
        errors = []
        asset_type = None
        asset_status = None
        if data.type_id:
            asset_type = (
                db_session.query(AssetTypeModel)
                .filter(AssetTypeModel.id == data.type_id)
                .first()
            )
            if not asset_type:
                errors.append(
                    {
                        "field": "assetType",
                        "error": f"Tipo de Ativo não existe. {asset_type}",
                    }
                )

        if data.status_id:
            asset_status = (
                db_session.query(AssetStatusModel)
                .filter(AssetStatusModel.id == data.status_id)
                .first()
            )
            if not asset_status:
                errors.append(
                    {
                        "field": "assetStatus",
                        "error": f"Situação de Ativo não existe. {asset_status}",
                    }
                )

        if len(errors) > 0:
            db_session.close()
            raise HTTPException(
                detail=errors,
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        return (
            asset_type,
            asset_status,
        )

    def __generate_registration_number(self, db_session: Session) -> str:
        """Generate new registration number"""
        last_asset = db_session.query(AssetModel).all()[-1]

        new_register_number = str(last_asset.id)

        return new_register_number.zfill(16 - len(new_register_number))

    def __get_asset_alert(self, asset: AssetModel) -> str:
        """Get asset alert"""
        message = ""
        low_count = 0
        medium_count = 0
        high_count = 0

        for maintenance in asset.maintenances:
            if maintenance.criticality_id == 1:
                low_count += 1
            elif maintenance.criticality_id == 2:
                medium_count += 1
            elif maintenance.criticality_id == 3:
                high_count += 1

        if high_count > 3:
            message = "Muitas manutenções críticas."
        elif medium_count > 5:
            message = "Muitas manutenções médias."
        elif low_count > 10:
            message = "Muitas manutenções leves."
        return message

    def serialize_asset(self, asset: AssetModel) -> AssetSerializerSchema:
        """Serialize asset"""
        last_maintenance = asset.maintenances[-1] if len(asset.maintenances) else None
        last_upgrade = asset.upgrades[-1] if len(asset.upgrades) else None

        return AssetSerializerSchema(
            id=asset.id,
            type=(
                AssetTypeSerializerSchema(**asset.type.__dict__) if asset.type else None
            ),
            status=(
                AssetStatusSerializerSchema(**asset.status.__dict__)
                if asset.status
                else None
            ),
            register_number=asset.register_number,
            description=asset.description,
            supplier=asset.supplier,
            assurance_date=asset.assurance_date,
            observations=asset.observations,
            discard_reason=asset.discard_reason,
            pattern=asset.pattern,
            operational_system=asset.operational_system,
            serial_number=asset.serial_number,
            imei=asset.imei,
            acquisition_date=(
                asset.acquisition_date.isoformat() if asset.acquisition_date else None
            ),
            value=asset.value,
            depreciation=asset.depreciation,
            ms_office=asset.ms_office,
            line_number=asset.line_number,
            operator=asset.operator,
            model=asset.model,
            accessories=asset.accessories,
            configuration=asset.configuration,
            quantity=asset.quantity,
            unit=asset.unit,
            by_agile=asset.by_agile,
            invoice_number=(asset.invoice.number if asset.invoice else None),
            maintenance_status=(
                last_maintenance.status.name
                if last_maintenance and last_maintenance.status.id != 3
                else "-"
            ),
            upgrade_status=(
                last_upgrade.status.name
                if last_upgrade and last_upgrade.status.id != 3
                else "-"
            ),
            alert=self.__get_asset_alert(asset),
        )

    def serialize_asset_type(
        self, asset_type: AssetTypeModel
    ) -> AssetTypeSerializerSchema:
        """Serialize asset type"""

        return AssetTypeSerializerSchema(**asset_type.__dict__)

    def serialize_asset_status(
        self, asset_status: AssetStatusModel
    ) -> AssetStatusSerializerSchema:
        """Serialize asset status"""

        return AssetStatusSerializerSchema(**asset_status.__dict__)

    def create_asset(
        self, data: NewAssetSchema, db_session: Session, authenticated_user: UserModel
    ) -> AssetSerializerSchema:
        """Creates new asset"""
        errors = []
        if (
            data.code
            and db_session.query(AssetModel)
            .filter(AssetModel.code == data.code)
            .first()
        ):
            errors.append({"field": "code", "error": "Este Código já existe."})
        if (
            data.register_number
            and db_session.query(AssetModel)
            .filter(AssetModel.register_number == data.register_number)
            .first()
        ):
            errors.append(
                {
                    "field": "registerNumber",
                    "error": "Este N° de Patrimônio já existe",
                }
            )

        if len(errors) > 0:
            db_session.close()
            raise HTTPException(
                detail=errors,
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        if data.register_number:
            register_number = data.register_number
        else:
            register_number = self.__generate_registration_number(db_session)

        (
            asset_type,
            asset_status,
        ) = self.__validate_nested(data, db_session)

        new_asset = AssetModel(
            code=data.code,
            register_number=register_number,
            description=data.description,
            supplier=data.supplier,
            assurance_date=data.assurance_date,
            observations=data.observations,
            discard_reason=data.discard_reason,
            pattern=data.pattern,
            operational_system=data.operational_system,
            serial_number=data.serial_number,
            imei=data.imei,
            acquisition_date=data.acquisition_date,
            value=data.value,
            ms_office=data.ms_office,
            line_number=data.line_number,
            operator=data.operator,
            model=data.model,
            accessories=data.accessories,
            configuration=data.configuration,
            quantity=data.quantity,
            unit=data.unit,
            active=True,
            by_agile=True,
        )

        new_asset.type = asset_type

        db_session.add(new_asset)
        db_session.commit()
        db_session.flush()

        self.update_asset_status(new_asset, asset_status, db_session)

        service_log.set_log(
            "lending",
            "asset",
            "Criação de Ativo",
            new_asset.id,
            authenticated_user,
            db_session,
        )
        logger.info("New Asset. %s", str(new_asset))

        return self.serialize_asset(new_asset)

    def update_asset(
        self,
        asset_id: int,
        data: UpdateAssetSchema,
        db_session: Session,
        authenticated_user: UserModel,
    ) -> AssetSerializerSchema:
        """Uptades an asset"""
        asset = self.__get_asset_or_404(asset_id, db_session)

        dict_data = data.model_dump()
        if asset.by_agile:
            for key, value in dict_data.items():
                if value is not None and key not in ["type_id", "status_id"]:
                    setattr(asset, key, value)

        else:
            fields_to_update = ["observations", "model", "line_number", "operator"]
            for field in fields_to_update:
                value = getattr(dict_data, field)
                if hasattr(dict_data, field) and value is not None:
                    setattr(asset, field, value)

        (
            asset_type,
            asset_status,
        ) = self.__validate_nested(data, db_session)

        if asset_type:
            asset.type = asset_type
        if asset_status:
            self.update_asset_status(asset, asset_status, db_session)

        db_session.add(asset)
        db_session.commit()
        db_session.flush()

        service_log.set_log(
            "lending",
            "asset",
            "Edição de Ativo",
            asset.id,
            authenticated_user,
            db_session,
        )
        logger.info("Updated Asset. %s", str(asset))
        return self.serialize_asset(asset)

    def inactivate_asset(
        self,
        asset_id: int,
        data: InactivateAssetSchema,
        db_session: Session,
        authenticated_user: UserModel,
    ) -> AssetSerializerSchema:
        """Uptades an asset"""
        asset = self.__get_asset_or_404(asset_id, db_session)

        asset.active = data.active

        db_session.add(asset)
        db_session.commit()
        db_session.flush()

        service_log.set_log(
            "lending",
            "asset",
            "Inativação de Ativo",
            asset.id,
            authenticated_user,
            db_session,
        )
        logger.info("Inactivate Asset. %s", str(asset))
        return self.serialize_asset(asset)

    def get_asset(self, asset_id: int, db_session: Session) -> AssetSerializerSchema:
        """Get an asset"""
        asset = self.__get_asset_or_404(asset_id, db_session)
        return self.serialize_asset(asset)

    def get_assets(
        self,
        db_session: Session,
        asset_filters: AssetFilter,
        ids: str = "",
        fields: str = "",
        page: int = 1,
        size: int = 50,
    ) -> Page[AssetSerializerSchema]:
        """Get assets list"""

        asset_list = asset_filters.filter(
            db_session.query(AssetModel)
            .outerjoin(AssetTypeModel)
            .outerjoin(AssetStatusModel)
        ).order_by(desc(AssetModel.id))

        if ids != "":
            list_ids = (
                [int(str_id) for str_id in ids.split(",")] if "," in ids else [int(ids)]
            )
            asset_list = (
                db_session.query(AssetModel)
                .filter(AssetModel.id.in_(list_ids))
                .union(asset_list)
            )

        params = Params(page=page, size=size)
        if fields == "":
            paginated = paginate(
                asset_list,
                params=params,
                transformer=lambda asset_list: [
                    self.serialize_asset(asset).model_dump(by_alias=True)
                    for asset in asset_list
                ],
            )
            return paginated

        list_fields = fields.split(",")
        paginated = paginate(
            asset_list,
            params=params,
            transformer=lambda asset_list: [
                self.serialize_asset(asset).model_dump(
                    include={*list_fields}, by_alias=True
                )
                for asset in asset_list
            ],
        )
        return paginated

    def get_asset_types(
        self,
        db_session: Session,
        filter_asset_type: AssetTypeFilter,
        fields: str = "",
    ) -> List[AssetTypeSerializerSchema]:
        """Get asset types list"""

        asset_type_list = filter_asset_type.filter(db_session.query(AssetTypeModel))

        if fields == "":
            return [
                self.serialize_asset_type(asset_type).model_dump(by_alias=True)
                for asset_type in asset_type_list
            ]

        list_fields = fields.split(",")
        return [
            self.serialize_asset_type(asset_type).model_dump(
                include={*list_fields}, by_alias=True
            )
            for asset_type in asset_type_list
        ]

    def get_asset_status(
        self,
        db_session: Session,
        filter_asset_status: AssetStatusFilter,
        fields: str = "",
    ) -> List[AssetTypeSerializerSchema]:
        """Get asset status list"""

        asset_status = filter_asset_status.filter(db_session.query(AssetStatusModel))

        if fields == "":
            return [
                self.serialize_asset_status(asset_status).model_dump(by_alias=True)
                for asset_status in asset_status
            ]

        list_fields = fields.split(",")
        return [
            self.serialize_asset_status(asset_status).model_dump(
                include={*list_fields}, by_alias=True
            )
            for asset_status in asset_status
        ]

    def get_asset_lending_history(
        self, asset_id: int, db_session: Session
    ) -> List[dict]:
        """Get an asset lending history"""
        asset = self.__get_asset_or_404(asset_id, db_session)

        historic_asset = (
            db_session.query(LendingModel)
            .filter(LendingModel.asset_id == asset.id, LendingModel.deleted.is_(False))
            .order_by(desc(LendingModel.id))
            .all()
        )

        historic_serialize = [
            LendingAssetHistorySerializerSchema(
                asset=h.asset.id,
                id=h.id,
                cost_center=CostCenterSerializerSchema(**h.cost_center.__dict__),
                document=h.document.id if h.document else None,
                document_revoke=h.document_revoke.id if h.document_revoke else None,
                employee=EmployeeShortSerializerSchema(
                    id=h.employee.id,
                    code=h.employee.code,
                    full_name=h.employee.full_name,
                    registration=h.employee.registration,
                ),
                glpi_number=h.glpi_number,
                number=h.number,
                observations=h.observations,
                project=h.project,
                revoke_signed_date=(
                    h.revoke_signed_date.strftime(DEFAULT_DATE_FORMAT)
                    if h.revoke_signed_date
                    else None
                ),
                signed_date=(
                    h.signed_date.strftime(DEFAULT_DATE_FORMAT)
                    if h.signed_date
                    else None
                ),
                status=h.status.name if h.status else None,
                witnesses=[witness.id for witness in h.witnesses],
                workload=h.workload.name if h.workload else "",
            ).model_dump(by_alias=True)
            for h in historic_asset
        ]

        return historic_serialize

    def update_asset_status(
        self,
        asset: AssetModel,
        asset_status: AssetStatusModel,
        db_session: Session,
        only_history: bool = False,
    ) -> None:
        """Update asset status"""
        if not only_history:
            asset.status = asset_status

            db_session.add(asset)
            db_session.commit()
            db_session.flush()

        historic = AssetStatusHistoricModel(
            asset_id=asset.id,
            status_id=asset_status.id,
        )

        db_session.add(historic)
        db_session.commit()
        db_session.flush()

    def __extract_data_from_row(
        self, row, header: List, record: dict, db_session: Session
    ) -> Union[dict, None]:
        for i, h in enumerate(header):
            value = row[i]
            key = self.MAP_COL_NAMES[h]
            if (
                key == "imei"
                and db_session.query(
                    db_session.query(AssetModel)
                    .filter(AssetModel.imei == value)
                    .exists()
                ).scalar()
            ):
                return {"error": f"IMEI já cadastrado: {value}"}
            if key == "register_number":
                if (
                    value
                    and db_session.query(
                        db_session.query(AssetModel)
                        .filter(AssetModel.register_number == value)
                        .exists()
                    ).scalar()
                ):
                    return {"error": f"N° de Patrimônio já cadastrado: {value}"}
                value = self.__generate_registration_number(db_session)
            if key == "invoice_id":
                invoice = (
                    db_session.query(InvoiceModel)
                    .filter(InvoiceModel.number == value)
                    .first()
                )
                if invoice:
                    record.update({"invoice_id": invoice.id})
                else:
                    new_invoice = InvoiceModel(number=value)
                    db_session.add(new_invoice)
                    db_session.commit()
                    db_session.flush()

                    record.update({"invoice_id": new_invoice.id})
            elif key == "type_id":
                asset_type = (
                    db_session.query(AssetTypeModel)
                    .filter(AssetTypeModel.name.ilike(value))
                    .first()
                )
                if not asset_type:
                    return {"error": f"Tipo de Ativo não existe: {value}"}
                record.update({key: asset_type.id})
            elif key == "status_id":
                asset_status = (
                    db_session.query(AssetStatusModel)
                    .filter(AssetStatusModel.name.ilike(value))
                    .first()
                )
                if not asset_status:
                    return {"error": f"Situação de Ativo não existe: {value}"}
                record.update({key: asset_status.id})
            else:
                record.update({key: value})
        return None

    async def upload_file_to_bulk_create(
        self, db_session: Session, file: UploadFile
    ) -> dict:
        """Upload file to bulk create"""
        file_bytes = await file.read()
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        sheet = workbook.active

        # Ler o cabeçalho e verificar se as colunas esperadas estão presentes
        header = [cell.value for cell in sheet[1]]
        if any(h not in self.COL_NAMES for h in header):
            missing_columns = [h for h in self.COL_NAMES if h not in header]
            return {"error": f"Colunas não existentes: {', '.join(missing_columns)}"}

        # Ler os dados
        new_assets = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {}
            error = self.__extract_data_from_row(row, header, record, db_session)
            if error:
                return error
            new_asset = AssetModel(**record, by_agile=True)
            new_assets.append(new_asset)

        db_session.bulk_save_objects(new_assets)
        db_session.commit()

        return {"message": "Arquivo enviado com sucesso."}
